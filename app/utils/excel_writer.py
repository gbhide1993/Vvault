import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import io


# 🎨 Colors
GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def write_answers(sheet_data, answers, rows):
    """
    Step 1: Populate dataframe (existing logic + new columns)
    Step 2: Write Excel using pandas
    Step 3: Re-open with openpyxl → apply colors
    """

    # -----------------------------
    # STEP 1: Populate DataFrames
    # -----------------------------
    for sheet_name, data in sheet_data.items():
        df = data["df"]

        # --- DOCX path: df is None, build from scratch ---
        if df is None:
            rows_data = []
            for row in rows:
                if row["sheet"] != sheet_name:
                    continue
                idx = row["index"]
                ans = answers[idx] if idx < len(answers) else {}
                confidence = ans.get("confidence", 0) if isinstance(ans, dict) else 0
                rows_data.append({
                    "Question": row["question"],
                    "Answer": ans.get("answer", "") if isinstance(ans, dict) else str(ans),
                    "Confidence": confidence,
                    "Source": ans.get("source", "") if isinstance(ans, dict) else "",
                    "Review Needed": "YES" if confidence < 70 else "NO",
                })
            data["df"] = pd.DataFrame(rows_data)
            continue

        # 🔍 Find answer column
        answer_col = None
        for col in df.columns:
            if col.lower() in ["answer", "response", "status"]:
                answer_col = col
                break

        if not answer_col:
            continue

        # 🔥 Fix dtype issue
        df[answer_col] = df[answer_col].astype(str)

        # 🔥 Add columns safely
        if "Confidence" not in df.columns:
            df["Confidence"] = ""

        if "Source" not in df.columns:
            df["Source"] = ""

        if "Review Needed" not in df.columns:
            df["Review Needed"] = ""

        df["Confidence"] = df["Confidence"].astype(str)
        df["Source"] = df["Source"].astype(str)
        df["Review Needed"] = df["Review Needed"].astype(str)

        # 🔁 Fill values
        for row in rows:
            if row["sheet"] != sheet_name:
                continue

            idx = row["index"]
            row_idx = row["row_idx"]

            answer_data = answers[idx]

            if isinstance(answer_data, dict):
                answer = answer_data.get("answer", "")
                confidence = answer_data.get("confidence", 0)
                confidence = float(confidence)

                # normalize if needed
                if confidence > 1:
                    confidence = confidence / 100
                source = answer_data.get("source", "")

                df.at[row_idx, answer_col] = answer
                df.at[row_idx, "Confidence"] = str(confidence)
                df.at[row_idx, "Source"] = source

                # 🔥 Review flag
                if float(confidence) < 0.70:
                    df.at[row_idx, "Review Needed"] = "YES"
                else:
                    df.at[row_idx, "Review Needed"] = "NO"

            else:
                df.at[row_idx, answer_col] = answer_data

        data["df"] = df

    # -----------------------------
    # STEP 2: Write Excel (pandas)
    # -----------------------------
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, data in sheet_data.items():
            data["df"].to_excel(writer, sheet_name=sheet_name, index=False)

    output.seek(0)

    # -----------------------------
    # STEP 3: Apply colors (openpyxl)
    # -----------------------------
    wb = load_workbook(output)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Find column indices
        headers = [cell.value for cell in ws[1]]

        try:
            answer_col_idx = headers.index(
                next(h for h in headers if str(h).lower() in ["answer", "response", "status"])
            ) + 1
        except StopIteration:
            continue

        if "Confidence" not in headers:
            continue

        confidence_col_idx = headers.index("Confidence") + 1

        # Apply colors row-wise
        for row in range(2, ws.max_row + 1):
            try:
                confidence_val = ws.cell(row=row, column=confidence_col_idx).value

                if confidence_val is None or confidence_val == "":
                    continue

                confidence_pct = float(confidence_val)

                # handle both formats (0.85 and 85)
                if confidence_pct <= 1:
                    confidence_pct *= 100

                if confidence_pct >= 80:
                    fill = GREEN
                elif 61 <= confidence_pct <= 79:
                    fill = YELLOW
                else:
                    fill = RED

                ws.cell(row=row, column=answer_col_idx).fill = fill

            except:
                continue

    # Save final output
    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    return final_output